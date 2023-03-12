import json
import openpyxl
from tempfile import NamedTemporaryFile
from openpyxl.styles import PatternFill, Border, Side, Font


def get_empty_crossword(data):
    name = 'e' + hash(data.__str__()).__str__()
    wb = openpyxl.Workbook()
    ws = wb.active
    sd = Side(border_style="medium", color="00000000")
    empty = Side(border_style="medium", color="00FFFFFF")

    size = len(data["matrix"])
    i = 0
    step = 0
    letter = 65
    fl = data["first_letters"]
    dots = []
    fl = data["first_letters"]
    for x in fl.keys():
        for st in range(3):
            dots.append(fl[x][st])
    print(dots)
    for elem in data["matrix"]:
        for j in range (size):
            let = chr(letter)
            ws.column_dimensions[let].border = Border(top=empty, bottom=empty, left=empty, right=empty)
            ws.row_dimensions[i].border = Border(top=empty, bottom=empty, left=empty, right=empty)
            if (elem[j] != "_"):
                str = f'{let}{i + 1}'
                ws[str].value = ' '
                ws[str].border = Border(top=sd, bottom=sd, left=sd, right=sd)
            d = 0
            while (d < len(dots)):
                if (i == dots[d + 1] and j == dots[d]):
                    str = f'{let}{i + 1}'
                    ws[str].value = dots[d + 2]
                d += 3
            ws.column_dimensions[let].width = 3
            letter += 1
            step += 1
            if (step % size == 0):
                letter = 65
                i += 1
    i = size + 1
    desc = data["descriptions"]
    wrd = ""
    dsc = ""
    k = 2
    for keys, vals in desc.items():
        ws.row_dimensions[i].border = Border(top=empty, bottom=empty, left=empty, right=empty)
        ws.row_dimensions[i + 1].border = Border(top=empty, bottom=empty, left=empty, right=empty)
        dsc = "".join(vals)
        adr_str = f'A{i}'
        dsc_str = f'A{i + 1}'
        ws[dsc_str] = dsc
        ws[adr_str] = dots[k]
        k += 3
        i += 2
    wb.save("empty.xlsx")
    return "empty.xlsx"
