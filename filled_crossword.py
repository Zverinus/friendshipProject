import json
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font


def get_filled_crossword(data):
    name = 'f' + hash(data.__str__()).__str__()
    wb = openpyxl.Workbook()
    ws = wb.active
    sd = Side(border_style="double", color="00000000")
    empty = Side(border_style="double", color="00FFFFFF")
    size = len(data["matrix"])
    i = 0
    step = 0
    letter = 65
    dots = []
    fl = data["first_letters"]
    for x in fl.keys():
        for st in range(3):
            dots.append(fl[x][st])
    for elem in data["matrix"]:
        for j in range (size):
            let = chr(letter)
            ws.column_dimensions[let].border = Border(top=empty, bottom=empty, left=empty, right=empty)
            ws.row_dimensions[i].border = Border(top=empty, bottom=empty, left=empty, right=empty)
            if (elem[j] != "_"):
                str = f'{let}{i + 1}'
                ws[str].value = elem[j]
                ws[str].border = Border(top=sd, bottom=sd, left=sd, right=sd)
            ws.column_dimensions[let].width = 2.5
            letter += 1
            step += 1
            if (step % size == 0):
                letter = 65
                i += 1
    desc = data["descriptions"]
    wrd = ""
    dsc = ""
    k = 2
    for keys, vals in desc.items():
        ws.row_dimensions[i].border = Border(top=empty, bottom=empty, left=empty, right=empty)
        wrd = "".join(keys)
        dsc = "".join(vals)
        wrd = "".join(keys)
        dsc = "".join(vals)
        adr_str = f'A{i}'
        wrd_str = f'B{i}'
        dsc_str = f'C{i}'
        ws[wrd_str] = wrd
        ws[dsc_str] = dsc
        ws[adr_str] = dots[k]
        k += 3
        i += 1
    wb.save("filled.xlsx")
    return "filled.xlsx"
